"""
Re-génère la colonne `Autres_albums_biblio` (BM Lyon) pour corriger la
pollution historique (albums d'autres artistes), SANS relancer tout le
pipeline `A_Recuperer`.

Ne re-scrape QUE le sous-étape « autres albums du même artiste »
(`_find_other_bm_lyon_albums`), et seulement pour les lignes dont
`Autres_albums_biblio` est déjà non vide (les seules potentiellement
polluées). Bénéficie du correctif de re-vérification d'auteur sur la fiche
détail, donc les albums d'autres artistes sont désormais écartés.

Fichiers mis à jour (périmètre « Resultats + Pipeline ») :
  - data/Pipeline/resultats_cotes_<playlist>.csv     (source)
  - data/Resultats/resultats_final_<playlist>.csv    (consultation)
  - data/Resultats/resultats_final_<playlist>.xlsx   (in-place, formatage préservé)

Chaque fichier est sauvegardé en `.bak` avant écriture.

Optimisation : 1 recherche BM Lyon par ARTISTE (pas par ligne). La liste
complète des albums de l'artiste est mise en cache, puis par ligne on retire
l'album de la ligne courante. ~1068 artistes ⇒ compter plusieurs heures.

Reprise : le cache par artiste est persisté sur disque
(`data/Pipeline/refresh_aab_cache_<playlist>.json`). Un Ctrl+C puis une
relance reprend là où on s'était arrêté (les artistes déjà scrapés sont
sautés). Supprimer ce fichier pour repartir de zéro.

Usage :
    cd sources/A_Recuperer
    uv run python refresh_autres_albums.py --dry-run        # estime, ne touche rien, pas de réseau
    uv run python refresh_autres_albums.py                  # applique (réseau, plusieurs heures)
    uv run python refresh_autres_albums.py --playlist Zen   # une seule playlist
    HEADLESS=false uv run python refresh_autres_albums.py   # navigateur visible (debug)
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from pathlib import Path

import pandas as pd

# Import des helpers du scraper (re-vérification d'auteur incluse).
sys.path.insert(0, str(Path(__file__).parent))
from utils.scraper import _find_other_bm_lyon_albums, _primary_artist  # noqa: E402
from utils.text_match import normalize  # noqa: E402

ROOT = Path(__file__).parent.parent.parent
PIPELINE_DIR = ROOT / "data" / "Pipeline"
RESULTATS_DIR = ROOT / "data" / "Resultats"

HEADLESS = os.environ.get("HEADLESS", "true").lower() == "true"

AAB_COL = "Autres_albums_biblio"
CACHE_FLUSH_EVERY = 5  # persiste le cache tous les N artistes (sécurité reprise)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def discover_playlists(only: str | None) -> list[str]:
    """Liste les playlists à traiter d'après les resultats_cotes_*.csv présents."""
    found = []
    for f in sorted(PIPELINE_DIR.glob("resultats_cotes_*.csv")):
        if ":" in f.name:  # ignore les *:Zone.Identifier de Windows
            continue
        name = f.stem.replace("resultats_cotes_", "")
        found.append(name)
    if only:
        if only not in found:
            print(f"Playlist '{only}' introuvable. Disponibles : {', '.join(found) or '(aucune)'}")
            return []
        return [only]
    return found


def subtract_album(full_list_str: str, own_album: str) -> str:
    """Retire de la liste « Titre - Cote, Titre - Cote » le segment dont le
    titre correspond à `own_album` (l'album propre de la ligne courante).

    Le titre est la partie avant le dernier ' - ' du segment (la cote est
    après). Comparaison normalisée (accents/casse)."""
    if not full_list_str:
        return ""
    own_norm = normalize(own_album)
    kept = []
    for seg in full_list_str.split(", "):
        seg = seg.strip()
        if not seg:
            continue
        title = seg.rsplit(" - ", 1)[0] if " - " in seg else seg
        if own_norm and normalize(title) == own_norm:
            continue
        kept.append(seg)
    return ", ".join(kept)


def load_cache(path: Path) -> dict[str, str]:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def save_cache(path: Path, cache: dict[str, str]) -> None:
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=0), encoding="utf-8")


def backup(path: Path) -> None:
    if path.exists():
        shutil.copy2(path, path.with_suffix(path.suffix + ".bak"))


# ---------------------------------------------------------------------------
# Cœur : traitement d'une playlist
# ---------------------------------------------------------------------------

def process_playlist(playlist: str, dry_run: bool, max_extra: int) -> None:
    cotes_path = PIPELINE_DIR / f"resultats_cotes_{playlist}.csv"
    final_csv = RESULTATS_DIR / f"resultats_final_{playlist}.csv"
    final_xlsx = RESULTATS_DIR / f"resultats_final_{playlist}.xlsx"
    cache_path = PIPELINE_DIR / f"refresh_aab_cache_{playlist}.json"

    df = pd.read_csv(cotes_path)
    if AAB_COL not in df.columns:
        print(f"[{playlist}] pas de colonne {AAB_COL} → rien à faire")
        return

    aab = df[AAB_COL].fillna("").astype(str).str.strip()
    affected_idx = df.index[aab != ""].tolist()

    # Regroupe par artiste (clé normalisée), garde un libellé brut représentatif
    jobs: dict[str, tuple[str, str]] = {}
    for i in affected_idx:
        raw = str(df.at[i, "Artist"])
        aq = _primary_artist(raw)
        key = normalize(aq)
        if key:
            jobs.setdefault(key, (aq, raw))

    print(f"[{playlist}] {len(affected_idx)} lignes à rafraîchir, "
          f"{len(jobs)} artistes uniques à re-scraper")

    if dry_run:
        print(f"[{playlist}] (dry-run) aucun réseau, aucun fichier modifié")
        return

    # --- Re-scrape (1 recherche par artiste, avec cache/reprise) ---
    cache = load_cache(cache_path)
    todo = [(k, v) for k, v in jobs.items() if k not in cache]
    print(f"[{playlist}] {len(cache)} artistes déjà en cache, {len(todo)} restants")

    if todo:
        from playwright.sync_api import sync_playwright

        debug_path = PIPELINE_DIR / "debug_selection.csv"
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=HEADLESS)
            context = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                locale="fr-FR",
            )
            page = context.new_page()
            try:
                for n, (key, (aq, raw)) in enumerate(todo, 1):
                    print(f"[{playlist}] ({n}/{len(todo)}) {aq!r}")
                    try:
                        full = _find_other_bm_lyon_albums(
                            page, aq, exclude_title="",
                            debug_path=debug_path, artist_brut=raw,
                            max_extra=max_extra,
                        )
                    except Exception as e:
                        print(f"   ! erreur sur {aq!r}: {e} → liste vide")
                        full = ""
                    cache[key] = full or ""
                    if n % CACHE_FLUSH_EVERY == 0:
                        save_cache(cache_path, cache)
            finally:
                save_cache(cache_path, cache)
                try:
                    context.close()
                    browser.close()
                except Exception:
                    pass

    # --- Recompose la valeur par ligne (liste complète - album de la ligne) ---
    new_by_key: dict[tuple[str, str], str] = {}
    for i in affected_idx:
        raw = str(df.at[i, "Artist"])
        album = str(df.at[i, "Album"])
        key = normalize(_primary_artist(raw))
        full = cache.get(key, "")
        new_val = subtract_album(full, album)
        df.at[i, AAB_COL] = new_val
        new_by_key[(raw, album)] = new_val

    n_now_empty = sum(1 for i in affected_idx if not str(df.at[i, AAB_COL]).strip())
    print(f"[{playlist}] recomposé : {len(affected_idx)} lignes, "
          f"{n_now_empty} désormais vides (artiste absent/sans autre album)")

    # --- Écriture Pipeline (source) ---
    backup(cotes_path)
    df.to_csv(cotes_path, index=False)
    print(f"[{playlist}] écrit {cotes_path.name} (sauvegarde .bak)")

    # --- Patch final CSV (consultation) ---
    if final_csv.exists():
        fdf = pd.read_csv(final_csv)
        if AAB_COL in fdf.columns and {"Artist_A_rechercher", "Album_A_rechercher"} <= set(fdf.columns):
            backup(final_csv)
            for j in fdf.index:
                k = (str(fdf.at[j, "Artist_A_rechercher"]), str(fdf.at[j, "Album_A_rechercher"]))
                if k in new_by_key:
                    fdf.at[j, AAB_COL] = new_by_key[k]
            fdf.to_csv(final_csv, index=False)
            print(f"[{playlist}] patché {final_csv.name} (sauvegarde .bak)")

    # --- Patch final XLSX in-place (préserve le formatage manuel) ---
    if final_xlsx.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(final_xlsx)
            ws = wb.active
            headers = {c.value: idx for idx, c in enumerate(ws[1], start=1)}
            if AAB_COL in headers and "Artist_A_rechercher" in headers and "Album_A_rechercher" in headers:
                backup(final_xlsx)
                aab_c = headers[AAB_COL]
                art_c = headers["Artist_A_rechercher"]
                alb_c = headers["Album_A_rechercher"]
                for r in ws.iter_rows(min_row=2):
                    k = (str(r[art_c - 1].value), str(r[alb_c - 1].value))
                    if k in new_by_key:
                        r[aab_c - 1].value = new_by_key[k] or None
                wb.save(final_xlsx)
                print(f"[{playlist}] patché {final_xlsx.name} in-place (sauvegarde .bak)")
            else:
                print(f"[{playlist}] colonnes attendues absentes du xlsx → xlsx non modifié")
        except Exception as e:
            print(f"[{playlist}] xlsx non patché ({e})")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Rafraîchit Autres_albums_biblio (BM Lyon) sans relancer le pipeline.")
    ap.add_argument("--dry-run", action="store_true", help="Compte seulement, aucun réseau, aucune écriture.")
    ap.add_argument("--playlist", default=None, help="Restreindre à une playlist (ex: Zen). Défaut : toutes.")
    ap.add_argument("--max-extra", type=int, default=8, help="Nb max d'autres albums listés par artiste (défaut 8).")
    args = ap.parse_args()

    playlists = discover_playlists(args.playlist)
    if not playlists:
        sys.exit(1)

    print(f"Playlists : {', '.join(playlists)}  |  mode : {'DRY-RUN' if args.dry_run else 'APPLY'}")
    for pl in playlists:
        process_playlist(pl, dry_run=args.dry_run, max_extra=args.max_extra)
    print("\nTerminé.")


if __name__ == "__main__":
    main()
